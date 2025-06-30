import discord
from discord import app_commands, ui, ButtonStyle, Embed, Color, Interaction, Member, Role, TextChannel, ChannelType
from discord.ext import commands
from db_handler import db
import config
import re
from datetime import datetime, timedelta
from io import BytesIO

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None

# --- 定数 ---
REPROCESS_EMOJI = '🔄'
DAYS_JP = ["月", "火", "水", "木", "金", "土", "日"]
DAYS_JP_FULL = ["月曜", "火曜", "水曜", "木曜", "金曜", "土曜", "日曜"]
DAYS_JP_ALL = DAYS_JP_FULL + DAYS_JP # "月曜"を先にチェックするため

# --- ★★★★★ ここから下のヘルパー関数を全て置き換えます ★★★★★ ---

def parse_schedule_message(message_content: str) -> list[dict] | None:
    lines = message_content.strip().split('\n')
    parsed_schedules = []
    day_map_rev = {day[0]: i for i, day in enumerate(DAYS_JP_ALL)}

    for line in lines:
        line = line.strip()
        if not line: continue
        
        days_to_process, rest_of_line = [], ""
        
        # 曜日範囲パターンを先にチェック
        range_match = re.match(r"([月火水木金土日])(?:曜)?\s*[~〜-]\s*([月火水木金土日])(?:曜)?(.*)", line)
        if range_match:
            start_day, end_day, rest = range_match.groups()
            start_idx, end_idx = day_map_rev.get(start_day), day_map_rev.get(end_day)
            if start_idx is not None and end_idx is not None and start_idx <= end_idx:
                days_to_process = DAYS_JP[start_idx : end_idx + 1]
                rest_of_line = rest.strip()
            else: continue
        else: # 単日パターン
            found_day = None
            for day_str in DAYS_JP_ALL:
                if line.startswith(day_str):
                    found_day = day_str[0]
                    rest_of_line = line[len(day_str):].strip()
                    break
            if found_day:
                days_to_process.append(found_day)
            else: continue

        # 時間と状態を解析
        time_str = rest_of_line
        status_jp = "参加"
        for s in ["一時参加", "参加", "休み", "無理", "不参加"]:
            if rest_of_line.endswith(s):
                status_jp = s
                time_str = rest_of_line[:-len(s)].strip()
                break
        
        status_en = "休み" if status_jp in ["休み", "無理", "不参加"] else status_jp
        time_str = "終日" if not time_str else time_str
        
        for day in days_to_process:
            parsed_schedules.append({"day": day, "time": time_str, "status": status_en})

    return parsed_schedules if parsed_schedules else None

def get_max_name_length(schedules: dict) -> int:
    max_len = 4
    for user_data in schedules.values():
        name = user_data.get("name", "")
        length = sum(2 if ord(char) > 255 else 1 for char in name)
        if length > max_len: max_len = length
    return max_len

def format_name(name: str, max_len: int) -> str:
    current_len = sum(2 if ord(char) > 255 else 1 for char in name)
    padding = " " * (max_len - current_len)
    return f"{name}{padding}"

def parse_time_range(time_str: str, default_start="20:00", default_end="24:00"):
    if not time_str or time_str == "終日": return (default_start, default_end)
    time_str = time_str.strip()
    # 全角チルダにも対応
    m = re.match(r"(\d{1,2}):(\d{2})\s*[~〜-]\s*(\d{1,2}):(\d{2})", time_str)
    if m: return (f"{int(m.group(1)):02}:{m.group(2)}", f"{int(m.group(3)):02}:{m.group(4)}")
    m = re.match(r"(\d{1,2}):(\d{2})\s*まで", time_str)
    if m: return (default_start, f"{int(m.group(1)):02}:{m.group(2)}")
    m = re.match(r"(\d{1,2}):(\d{2})\s*[~〜から]", time_str)
    if m: return (f"{int(m.group(1)):02}:{m.group(2)}", default_end)
    return (default_start, default_end)

def time_range_blocks(start="20:00", end="24:00", interval_min=30):
    try:
        start_dt = datetime.strptime(start, "%H:%M")
        end_dt = datetime.strptime(end, "%H:%M") if end != "24:00" else datetime.strptime("23:59", "%H:%M") + timedelta(minutes=1)
    except ValueError: start_dt, end_dt = datetime.strptime("20:00", "%H:%M"), datetime.strptime("23:59", "%H:%M") + timedelta(minutes=1)
    blocks, t = [], start_dt
    while t < end_dt:
        blocks.append((t.strftime("%H:%M"), (t + timedelta(minutes=interval_min)).strftime("%H:%M")))
        t += timedelta(minutes=interval_min)
    return blocks

def is_in_timeblock(tblock, user_start, user_end):
    try:
        s0, s1 = datetime.strptime(tblock[0], "%H:%M"), datetime.strptime(tblock[1], "%H:%M")
        r0, r1 = datetime.strptime(user_start, "%H:%M"), datetime.strptime(user_end, "%H:%M") if user_end != "24:00" else datetime.strptime("23:59", "%H:%M") + timedelta(minutes=1)
        return max(s0, r0) < min(s1, r1)
    except ValueError: return False
        
# --- Cog本体 ---
class ShiftCog(commands.Cog):
    """週間の活動予定（シフト）を管理する機能"""
    help_category = "シフト管理"
    help_description = "メンバーの週間活動予定をスレッドで収集し、一覧化します。"
    command_helps = {
        "shift create": "指定した対象の予定調整スreadを作成します。",
        "shift create_all": "クランメンバー全員の予定調整スレッドを一斉に作成します。",
        "shift export": "全員分の予定を集計し、シフト表としてチャンネルに投稿します。",
        "shift cleanup": "作成した全ての予定調整スレッドを一斉にアーカイブします。",
        "shift export_timeline_excel": "さらに詳細な表をエクセルで作成します。",
    }

    def __init__(self, bot: commands.Bot):
        self.bot = bot

    async def cog_app_command_error(self, interaction: discord.Interaction, error: app_commands.AppCommandError):
        print(f"Cog 'ShiftCog' でエラー: {error}")
        import traceback; traceback.print_exc()
        if not interaction.response.is_done():
            await interaction.response.send_message("❌ 処理中にエラーが発生しました。", ephemeral=True)

    async def _process_schedule_message(self, message: discord.Message):
        if message.author.bot or not isinstance(message.channel, discord.Thread): return
        schedules = db.get("shift_schedules", {})
        thread_id_str, user_id_str = str(message.channel.id), None
        for uid, udata in schedules.items():
            if udata.get("thread_id") == thread_id_str:
                user_id_str = uid
                break
        if not user_id_str or user_id_str != str(message.author.id): return
        parsed_list = parse_schedule_message(message.content)
        if parsed_list is None:
            error_message = (f"{message.author.mention} 書き方が違うようです！\n"
                             "基本の形: `曜日 時間 状態` で、1行ずつ改行して入力してくださいね。")
            try: await message.reply(error_message, delete_after=15)
            except discord.Forbidden: pass
            return
        for parsed in parsed_list:
            day_key = f"day_{parsed['day']}"
            schedules[user_id_str].setdefault("schedule", {})[day_key] = f"{parsed['time']} ({parsed['status']})"
        db.set("shift_schedules", schedules)
        try: await message.add_reaction("✅")
        except discord.Forbidden: print(f"ERROR: リアクション付与権限がありません in {message.channel.name}")

    @commands.Cog.listener()
    async def on_message(self, message: discord.Message):
        await self._process_schedule_message(message)

    @commands.Cog.listener()
    async def on_raw_reaction_add(self, payload: discord.RawReactionActionEvent):
        if payload.user_id == self.bot.user.id or str(payload.emoji) != REPROCESS_EMOJI: return
        try:
            channel = await self.bot.fetch_channel(payload.channel_id)
            message = await channel.fetch_message(payload.message_id)
            await self._process_schedule_message(message)
            await message.remove_reaction(payload.emoji, self.bot.user)
            await message.remove_reaction(payload.emoji, discord.Object(id=payload.user_id))
        except (discord.NotFound, discord.Forbidden): pass

    shift = app_commands.Group(name="shift", description="週間活動予定（シフト）の管理", guild_only=True)

    # ★★★★★ ここが修正箇所 ★★★★★
    async def _create_schedule_thread(self, channel: TextChannel, member: Member):
        schedules = db.get("shift_schedules", {})
        user_id_str = str(member.id)
        if user_id_str in schedules and schedules[user_id_str].get("thread_id"):
            return None, "既にスレッドが存在します"
        try:
            thread = await channel.create_thread(name=f"週間予定 - {member.display_name}", type=ChannelType.private_thread)
            await thread.add_user(member)
            schedules.setdefault(user_id_str, {})["thread_id"] = str(thread.id)
            schedules[user_id_str]["name"] = member.display_name
            db.set("shift_schedules", schedules)

            # スタッフロールを取得してメンションを作成
            staff_mention = ""
            if config.STAFF_ROLE_ID:
                staff_role = channel.guild.get_role(config.STAFF_ROLE_ID)
                if staff_role:
                    staff_mention = f"{staff_role.mention} "

            initial_message_content = f"{staff_mention}{member.mention}さんのシフト調整スレッドが作成されました。"

            initial_embed = Embed(title="週間活動予定を教えてください", color=Color.blue(), description=(
                "このスレッドに、今週の活動予定を書き込んでください。\n\n"
                "--- \n"
                "**【基本の書き方】**\n"
                "`曜日 時間 状態`\n\n"
                "`曜日`：月, 火, 水, 木, 金, 土, 日\n"
                "`時間`：「終日」または「21:00~23:30」「22時まで」など\n"
                "`状態`：「参加」「一時参加」「休み」など\n\n"
                "--- \n"
                "**【入力例 (複数行OK)】**\n"
                "```\n"
                "月曜 終日 参加\n"
                "火曜 21:30~23:00 参加\n"
                "水曜 22時まで 一時参加\n"
                "木曜 休み\n"
                "```\n"
                "--- \n"
                "予定を書き込むと、私が✅リアクションで確認の合図を送ります。"
            ))

            await thread.send(content=initial_message_content, embed=initial_embed)
            return thread, None
        except discord.Forbidden: return None, "スレッドの作成権限がありません。"
        except Exception as e: return None, f"不明なエラー: {e}"

    @shift.command(name="create", description="指定した対象の予定調整スレッドを作成します。")
    @app_commands.checks.has_permissions(manage_threads=True)
    @app_commands.describe(member="スレッドを作成するメンバー", role="スレッドを作成するロール")
    async def create(self, interaction: Interaction, member: Member = None, role: Role = None):
        if not member and not role: return await interaction.response.send_message("メンバーまたはロールのいずれか一方を指定してください。", ephemeral=True)
        await interaction.response.defer(ephemeral=True)
        targets = list(dict.fromkeys(([member] if member else []) + (role.members if role else [])))
        success_count, fail_count, skip_count, error_messages = 0, 0, 0, []
        for m in targets:
            if m.bot: continue
            thread, error = await self._create_schedule_thread(interaction.channel, m)
            if thread: success_count += 1
            elif "既に" in (error or ""): skip_count += 1
            else: fail_count += 1; error_messages.append(error) if error not in error_messages else None
        await interaction.followup.send(f"スレッド作成完了。\n✅ 成功: {success_count}件\n⏩ スキップ: {skip_count}件\n❌ 失敗: {fail_count}件\n{', '.join(error_messages)}", ephemeral=True)

    @shift.command(name="create_all", description="クランメンバー全員の予定調整スレッドを一斉に作成します。")
    @app_commands.checks.has_permissions(manage_threads=True)
    async def create_all(self, interaction: Interaction):
        if not config.CLAN_MEMBER_ROLE_ID: return await interaction.response.send_message("`CLAN_MEMBER_ROLE_ID`が設定されていません。", ephemeral=True)
        clan_member_role = interaction.guild.get_role(config.CLAN_MEMBER_ROLE_ID)
        if not clan_member_role: return await interaction.response.send_message("クランメンバーロールが見つかりません。", ephemeral=True)
        await self.create.callback(self, interaction, member=None, role=clan_member_role)

    @shift.command(name="export", description="全員分の予定を集計し、シフト表としてチャンネルに投稿します。")
    @app_commands.checks.has_permissions(manage_threads=True)
    async def export(self, interaction: Interaction):
        await interaction.response.defer()
        schedules = db.get("shift_schedules", {})
        if not schedules: return await interaction.followup.send("スケジュールデータがありません。")
        days = ["月", "火", "水", "木", "金", "土", "日"]
        max_name_len = get_max_name_length(schedules)
        header_name = format_name("名前", max_name_len)
        header = f"| {header_name} | {' | '.join(days)} |"
        separator = f"| :{'-' * max_name_len}: |{':---:|' * len(days)}"
        lines = [header, separator]
        for user_data in schedules.values():
            name = user_data.get("name", "不明")
            formatted_name = format_name(name, max_name_len)
            schedule = user_data.get("schedule", {})
            row = f"| {formatted_name} |"
            for day_jp in days:
                day_key = f"day_{day_jp}"; entry = schedule.get(day_key, "未")
                cell = "✅" if "参加" in entry else ("🕒" if "一時" in entry else ("❌" if "休み" in entry else "❔"))
                row += f" {cell} |"
            lines.append(row)
        output_str = "```\n" + "\n".join(lines) + "\n```"
        embed = Embed(title="週間活動シフト表", description=output_str, color=Color.blue(), timestamp=datetime.now())
        await interaction.followup.send(embed=embed)

    @shift.command(name="export_excel", description="全員分の予定を集計し、Excelファイルとして出力します。")
    @app_commands.checks.has_permissions(manage_threads=True)
    async def export_excel(self, interaction: Interaction):
        # (このコマンドの中身は変更なし)
        await interaction.response.defer(ephemeral=True)
        schedules = db.get("shift_schedules", {})
        if not schedules: return await interaction.followup.send("スケジュールデータがありません。", ephemeral=True)
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            from io import BytesIO
            wb = openpyxl.Workbook(); ws = wb.active; ws.title = "週間シフト表"
            days = ["月", "火", "水", "木", "金", "土", "日"]; header = ["名前"] + days; ws.append(header)
            header_font = Font(bold=True, color="FFFFFF"); header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            header_alignment = Alignment(horizontal='center', vertical='center')
            for cell in ws["1:1"]: cell.font = header_font; cell.fill = header_fill; cell.alignment = header_alignment
            status_colors = { "参加": PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"), "一時参加": PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid"), "休み": PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid"), "未": PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")}
            status_alignment = Alignment(horizontal='center', vertical='center')
            for user_data in schedules.values():
                name = user_data.get("name", "不明"); schedule = user_data.get("schedule", {}); row_data = [name]
                for day_jp in days:
                    day_key = f"day_{day_jp}"; entry_with_time = schedule.get(day_key, "未")
                    status = "参加" if "参加" in entry_with_time else ("一時参加" if "一時" in entry_with_time else ("休み" if "休み" in entry_with_time else "未"))
                    row_data.append(status)
                ws.append(row_data)
                row_index = ws.max_row
                for col_index, status_value in enumerate(row_data):
                    if col_index > 0:
                        cell = ws.cell(row=row_index, column=col_index + 1); cell.alignment = status_alignment
                        if status_value in status_colors: cell.fill = status_colors.get(status_value)
            for col_idx, col in enumerate(ws.columns, 1):
                max_length = 0; column_letter = col[0].column_letter
                if column_letter == 'A': ws.column_dimensions[column_letter].width = get_max_name_length(schedules) * 1.2 + 2; continue
                for cell in col:
                    if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
                ws.column_dimensions[column_letter].width = max_length + 2
            virtual_workbook = BytesIO(); wb.save(virtual_workbook); virtual_workbook.seek(0)
            file = discord.File(fp=virtual_workbook, filename=f"shift_{datetime.now().strftime('%Y%m%d')}.xlsx")
            await interaction.followup.send("✅ シフト表のExcelファイルを作成しました。", file=file, ephemeral=True)
        except ImportError: await interaction.followup.send("❌ `openpyxl`ライブラリがインストールされていません。", ephemeral=True)
        except Exception as e: await interaction.followup.send(f"❌ Excelファイルの作成中にエラーが発生しました: {e}", ephemeral=True)

    @shift.command(name="cleanup", description="作成した全ての予定調整スレッドを一斉にアーカイブ（削除）します。")
    @app_commands.checks.has_permissions(manage_threads=True)
    async def cleanup(self, interaction: Interaction):
        # (このコマンドの中身は変更なし)
        await interaction.response.defer(ephemeral=True)
        schedules = db.get("shift_schedules", {})
        if not schedules: return await interaction.followup.send("クリーンアップ対象のスレッドはありません。", ephemeral=True)
        archived_count, failed_count = 0, 0
        for user_id in list(schedules.keys()):
            thread_id = schedules[user_id].get("thread_id")
            if thread_id:
                try:
                    thread = await self.bot.fetch_channel(int(thread_id)); await thread.edit(archived=True, locked=True); archived_count += 1
                except (discord.NotFound, discord.Forbidden): failed_count += 1
            del schedules[user_id]
        db.set("shift_schedules", schedules)
        await interaction.followup.send(f"クリーンアップ完了。\n✅ アーカイブ成功: {archived_count}件\n❌ 失敗: {failed_count}件", ephemeral=True)

# cogs/shift.py の ShiftCog クラス内に追記

    @shift.command(name="export_day", description="指定した曜日の詳細なタイムライン形式のExcelファイルを出力します。")
    @app_commands.checks.has_permissions(manage_threads=True)
    @app_commands.describe(day="出力したい曜日を選択してください。")
    @app_commands.choices(day=[
        app_commands.Choice(name="月曜日", value="月"),
        app_commands.Choice(name="火曜日", value="火"),
        app_commands.Choice(name="水曜日", value="水"),
        app_commands.Choice(name="木曜日", value="木"),
        app_commands.Choice(name="金曜日", value="金"),
        app_commands.Choice(name="土曜日", value="土"),
        app_commands.Choice(name="日曜日", value="日"),
    ])
    async def export_day_excel(self, interaction: Interaction, day: app_commands.Choice[str]):
        await interaction.response.defer(ephemeral=True)

        schedules = db.get("shift_schedules", {})
        if not schedules:
            return await interaction.followup.send("スケジュールデータがありません。", ephemeral=True)

        try:
            # 必要なライブラリをインポート
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            from io import BytesIO

            # Excelブックとシートを作成
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"{day.name}のシフト"

            # 色の定義
            fills = {
                "参加": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),  # 緑
                "一時参加": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"), # 黄
                "休み": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"), # 赤
                "未定": PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),    # グレー
            }
            center_alignment = Alignment(horizontal='center', vertical='center')

            # ヘッダー行を作成 (名前, 20:00, 20:30, ...)
            time_blocks = time_range_blocks("20:00", "24:00", 30)
            header = ["名前"] + [block[0] for block in time_blocks]
            ws.append(header)
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = center_alignment
            
            # メンバーごとの行を作成
            for user_id, user_data in schedules.items():
                name = user_data.get("name", f"ID:{user_id}")
                schedule_for_day = user_data.get("schedule", {}).get(f"day_{day.value}", "未定") # day.valueは "月"など
                
                row_values = [name]

                # "(参加)" のような部分から状態を抽出
                match = re.search(r"\((.+?)\)$", schedule_for_day)
                status = match.group(1) if match else "未定"
                time_str = schedule_for_day.replace(f"({status})", "").strip()
                user_start, user_end = parse_time_range(time_str)

                # 各時間ブロックのセルを埋める
                for t_block in time_blocks:
                    cell_status = "未定"
                    if status == "休み":
                        cell_status = "休み"
                    elif status in ["参加", "一時参加"]:
                        if is_in_timeblock(t_block, user_start, user_end):
                            cell_status = status
                    
                    row_values.append(cell_status)
                
                ws.append(row_values)
                
                # セルに色を付ける
                row_index = ws.max_row
                for col_index, status_value in enumerate(row_values[1:], 2):
                    cell = ws.cell(row=row_index, column=col_index)
                    cell.fill = fills.get(status_value, fills["未定"])
            
            # 列幅を調整
            ws.column_dimensions['A'].width = get_max_name_length(schedules) * 1.2 + 4
            for i in range(2, len(header) + 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 10

            # メモリ上でファイルを保存
            virtual_workbook = BytesIO()
            wb.save(virtual_workbook)
            virtual_workbook.seek(0)
            
            file = discord.File(fp=virtual_workbook, filename=f"shift_{day.name}_{datetime.now().strftime('%Y%m%d')}.xlsx")
            await interaction.followup.send(f"✅ {day.name}のタイムライン形式Excelシフト表を作成しました。", file=file, ephemeral=True)

        except ImportError:
            await interaction.followup.send("❌ `openpyxl`ライブラリがインストールされていません。", ephemeral=True)
        except Exception as e:
            print(f"Excel作成エラー: {e}")
            import traceback
            traceback.print_exc()
            await interaction.followup.send(f"❌ Excelファイルの作成中にエラーが発生しました: {e}", ephemeral=True)
            
    @commands.Cog.listener()
    async def on_message(self, message: discord.Message):
        """スレッド内の書き込みを監視し、予定を自動で記録する"""
        if message.author.bot or not isinstance(message.channel, discord.Thread): return

        schedules = db.get("shift_schedules", {})
        thread_id_str = str(message.channel.id)
        user_id_str = next((uid for uid, udata in schedules.items() if udata.get("thread_id") == thread_id_str), None)
        
        if not user_id_str or user_id_str != str(message.author.id): return

        # ★★★ ここからデバッグログ付き ★★★
        print(f"\n\n--- on_message TRIGGERED by {message.author.display_name} ---")
        
        parsed_list = parse_schedule_message(message.content)
        
        if parsed_list is None:
            print(" -> 解析結果がNoneのため、エラーメッセージを返信します。")
            error_message = (f"{message.author.mention} 書き方が違うようです！\n"
                             "基本の形: `曜日 時間 状態` で入力してくださいね。")
            try: await message.reply(error_message, delete_after=20)
            except: pass
            return

        print(f" -> DB更新対象: {len(parsed_list)}件")
        for parsed in parsed_list:
            day_key = f"day_{parsed['day']}"
            schedules[user_id_str].setdefault("schedule", {})[day_key] = f"{parsed['time']} ({parsed['status']})"
        
        db.set("shift_schedules", schedules)
        print(" -> DB更新完了。")
        try:
            await message.add_reaction("✅")
            print(" -> リアクション付与完了。")
        except discord.Forbidden:
            print(f"ERROR: リアクション付与の権限がありません in {message.channel.name}")
        print("--- on_message FINISHED ---")

async def setup(bot: commands.Bot):
    await bot.add_cog(ShiftCog(bot))
